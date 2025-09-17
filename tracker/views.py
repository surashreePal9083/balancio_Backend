from django.contrib.auth import authenticate, login, logout
from django.db.models import Q, Sum
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status, generics, filters
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework.request import Request
from rest_framework.pagination import PageNumberPagination
from rest_framework.views import APIView
from datetime import datetime, timedelta
from decimal import Decimal
import json
import csv
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
from rest_framework_simplejwt.tokens import RefreshToken, AccessToken
from rest_framework_simplejwt.token_blacklist.models import BlacklistedToken, OutstandingToken
from rest_framework_simplejwt.exceptions import TokenError, InvalidToken
from django.contrib.auth import get_user_model

from .models import User, Transaction, Category, Notification, Budget, BudgetAlert, MonthlyReport
from .serializers import (
    UserSerializer, UserProfileSerializer, UserSettingsSerializer, ChangePasswordSerializer,
    LoginSerializer, RegisterSerializer, CategorySerializer, TransactionSerializer,
    BudgetSerializer, BudgetAlertSerializer, NotificationSerializer, MonthlyReportSerializer
)

# =============================================================================
# AUTHENTICATION ENDPOINTS
# =============================================================================

@swagger_auto_schema(
    method='post',
    operation_description="Authenticate a user with email and password",
    operation_summary="User login",
    tags=['Authentication'],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'email': openapi.Schema(type=openapi.TYPE_STRING, description='Email address'),
            'password': openapi.Schema(type=openapi.TYPE_STRING, description='Password')
        },
        required=['email', 'password']
    ),
    responses={
        200: openapi.Response(description="Login successful"),
        400: openapi.Response(description="Bad request"),
        401: openapi.Response(description="Invalid credentials")
    }
)
@api_view(['POST'])
@permission_classes([AllowAny])
def auth_login(request: Request) -> Response:
    """User login endpoint"""
    serializer = LoginSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.validated_data['user']
        refresh = RefreshToken.for_user(user)
        
        return Response({
            'message': 'Login successful!',
            'user': {
                'id': user.pk,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email
            },
            'access': str(refresh.access_token),
            'refresh': str(refresh)
        })
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@swagger_auto_schema(
    method='post',
    operation_description="Register a new user account",
    operation_summary="User registration",
    tags=['Authentication'],
    request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'first_name': openapi.Schema(type=openapi.TYPE_STRING),
            'last_name': openapi.Schema(type=openapi.TYPE_STRING),
            'email': openapi.Schema(type=openapi.TYPE_STRING),
            'password': openapi.Schema(type=openapi.TYPE_STRING)
        },
        required=['first_name', 'last_name', 'email', 'password']
    ),
    responses={
        201: openapi.Response(description="User created successfully"),
        400: openapi.Response(description="Bad request")
    }
)
@api_view(['POST'])
@permission_classes([AllowAny])
def auth_signup(request: Request) -> Response:
    """User registration endpoint"""
    serializer = RegisterSerializer(data=request.data)
    if serializer.is_valid():
        user = serializer.save()
        return Response({
            'message': 'User created successfully!',
            'user': {
                'id': user.pk,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'email': user.email
            }
        }, status=status.HTTP_201_CREATED)
    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([AllowAny])
def auth_refresh(request: Request) -> Response:
    """Refresh access token using refresh token"""
    try:
        refresh_token = request.data.get('refresh')
        if not refresh_token:
            return Response({
                'error': 'Refresh token is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Validate and refresh the token
            refresh = RefreshToken(refresh_token)
            
            # Get user from token payload
            user_id = refresh.payload.get('user_id')
            if not user_id:
                return Response({
                    'error': 'Invalid token payload'
                }, status=status.HTTP_401_UNAUTHORIZED)
            
            # Get user instance
            User = get_user_model()
            try:
                user = User.objects.get(id=user_id)
            except User.DoesNotExist:
                return Response({
                    'error': 'User not found'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # Generate new tokens
            new_refresh = RefreshToken.for_user(user)
            
            return Response({
                'access': str(new_refresh.access_token),
                'refresh': str(new_refresh),
                'message': 'Token refreshed successfully'
            })
            
        except TokenError as e:
            return Response({
                'error': 'Invalid or expired refresh token',
                'details': str(e)
            }, status=status.HTTP_401_UNAUTHORIZED)
            
    except Exception as e:
        return Response({
            'error': f'Token refresh failed: {str(e)}'
        }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([AllowAny])
def auth_verify(request: Request) -> Response:
    """Verify if access token is valid"""
    try:
        token = request.data.get('token')
        if not token:
            return Response({
                'error': 'Token is required'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Validate the access token
            validated_token = AccessToken(token)
            return Response({
                'valid': True,
                'user_id': validated_token.get('user_id'),
                'exp': validated_token.get('exp'),
                'message': 'Token is valid'
            })
            
        except TokenError as e:
            return Response({
                'valid': False,
                'error': 'Invalid or expired token',
                'details': str(e)
            }, status=status.HTTP_401_UNAUTHORIZED)
            
    except Exception as e:
        return Response({
            'error': f'Token verification failed: {str(e)}'
        }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def auth_logout(request: Request) -> Response:
    """User logout endpoint with token blacklisting"""
    try:
        # Get refresh token from request
        refresh_token = request.data.get('refresh')
        
        if refresh_token:
            try:
                # Blacklist the refresh token
                token = RefreshToken(refresh_token)
                token.blacklist()
                
                return Response({
                    'message': 'Logout successful! Token has been blacklisted.'
                })
            except TokenError as e:
                return Response({
                    'message': 'Logout completed, but token was already invalid.',
                    'details': str(e)
                })
        else:
            # If no refresh token provided, just log out the session
            logout(request)
            return Response({
                'message': 'Logout successful! (No token to blacklist)'
            })
            
    except Exception as e:
        return Response({
            'error': f'Logout failed: {str(e)}'
        }, status=status.HTTP_400_BAD_REQUEST)

# =============================================================================
# USER PROFILE AND SETTINGS ENDPOINTS
# =============================================================================

@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def users_profile(request: Request) -> Response:
    """Get or update user profile"""
    if request.method == 'GET':
        serializer = UserProfileSerializer(request.user)
        response_data = serializer.data
        
        # Add additional fields for frontend compatibility
        created_at_field = getattr(request.user, 'created_at', None)
        updated_at_field = getattr(request.user, 'updated_at', None)
        
        response_data.update({
            'firstName': getattr(request.user, 'first_name', ''),
            'lastName': getattr(request.user, 'last_name', ''),
            'name': f"{getattr(request.user, 'first_name', '')} {getattr(request.user, 'last_name', '')}".strip(),
            'avatar': getattr(request.user, 'profile_picture', None),
            'createdAt': created_at_field.isoformat() if created_at_field else None,
            'updatedAt': updated_at_field.isoformat() if updated_at_field else None,
            'settings': {
                'emailNotifications': getattr(request.user, 'email_notifications', True),
                'budgetAlerts': getattr(request.user, 'budget_alerts', True),
                'monthlyReports': getattr(request.user, 'monthly_reports', True),
                'reportFormat': getattr(request.user, 'preferred_currency', 'excel'),
                'twoFactorEnabled': False  # Placeholder for 2FA
            }
        })
        
        return Response(response_data)
    
    elif request.method == 'PUT':
        print(f"Profile update request data: {request.data}")  # Debug log
        
        # Handle different naming conventions from frontend
        update_data = request.data.copy()
        
        # Map frontend field names to backend field names
        if 'firstName' in update_data:
            update_data['first_name'] = update_data.pop('firstName')
        if 'lastName' in update_data:
            update_data['last_name'] = update_data.pop('lastName')
        if 'name' in update_data:
            # Split name into first_name and last_name
            name_parts = update_data['name'].split(' ', 1)
            update_data['first_name'] = name_parts[0] if name_parts else ''
            update_data['last_name'] = name_parts[1] if len(name_parts) > 1 else ''
            update_data.pop('name')
        
        print(f"Processed update data: {update_data}")  # Debug log
        
        serializer = UserProfileSerializer(request.user, data=update_data, partial=True)
        if serializer.is_valid():
            user = serializer.save()
            
            # Return updated user data in frontend-compatible format
            created_at_field = getattr(user, 'created_at', None)
            updated_at_field = getattr(user, 'updated_at', None)
            
            response_data = serializer.data
            response_data.update({
                'firstName': getattr(user, 'first_name', ''),
                'lastName': getattr(user, 'last_name', ''),
                'name': f"{getattr(user, 'first_name', '')} {getattr(user, 'last_name', '')}".strip(),
                'avatar': getattr(user, 'profile_picture', None),
                'createdAt': created_at_field.isoformat() if created_at_field else None,
                'updatedAt': updated_at_field.isoformat() if updated_at_field else None,
                'settings': {
                    'emailNotifications': getattr(user, 'email_notifications', True),
                    'budgetAlerts': getattr(user, 'budget_alerts', True),
                    'monthlyReports': getattr(user, 'monthly_reports', True),
                    'reportFormat': getattr(user, 'preferred_currency', 'excel'),
                    'twoFactorEnabled': False
                }
            })
            
            return Response({
                'message': 'Profile updated successfully!',
                'user': response_data
            })
        else:
            print(f"Serializer errors: {serializer.errors}")  # Debug log
            return Response({
                'error': 'Validation failed',
                'details': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
    
    return Response({'error': 'Method not allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

@api_view(['PUT'])
@permission_classes([IsAuthenticated])
def users_change_password(request: Request) -> Response:
    """Change user password"""
    print(f"Password change request data: {request.data}")  # Debug log
    
    # Handle different field names from frontend
    data = request.data.copy()
    if 'currentPassword' in data:
        data['old_password'] = data.pop('currentPassword')
    if 'newPassword' in data:
        data['new_password'] = data.pop('newPassword')
    if 'confirmPassword' in data:
        data['confirm_password'] = data.pop('confirmPassword')
    
    serializer = ChangePasswordSerializer(data=data)
    if serializer.is_valid():
        user = request.user
        
        # Check old password
        if not user.check_password(serializer.validated_data['old_password']):
            return Response({
                'error': 'Old password is incorrect',
                'message': 'Current password is incorrect'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Set new password
        user.set_password(serializer.validated_data['new_password'])
        user.save()
        
        return Response({
            'message': 'Password changed successfully!',
            'success': True
        })
    else:
        print(f"Password change serializer errors: {serializer.errors}")  # Debug log
        return Response({
            'error': 'Validation failed',
            'details': serializer.errors,
            'message': 'Password change failed'
        }, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def users_settings(request: Request) -> Response:
    """Get or update user settings"""
    if request.method == 'GET':
        serializer = UserSettingsSerializer(request.user)
        response_data = serializer.data
        
        # Add frontend-compatible settings format
        response_data['settings'] = {
            'emailNotifications': getattr(request.user, 'email_notifications', True),
            'budgetAlerts': getattr(request.user, 'budget_alerts', True),
            'monthlyReports': getattr(request.user, 'monthly_reports', True),
            'reportFormat': getattr(request.user, 'preferred_currency', 'excel'),
            'twoFactorEnabled': False
        }
        
        return Response(response_data)
    
    elif request.method == 'PUT':
        print(f"Settings update request data: {request.data}")  # Debug log
        
        # Handle frontend settings format
        update_data = {}
        
        # Map frontend settings to backend fields
        if 'emailNotifications' in request.data:
            update_data['email_notifications'] = request.data['emailNotifications']
        if 'budgetAlerts' in request.data:
            update_data['budget_alerts'] = request.data['budgetAlerts']
        if 'monthlyReports' in request.data:
            update_data['monthly_reports'] = request.data['monthlyReports']
        if 'reportFormat' in request.data:
            update_data['preferred_currency'] = request.data['reportFormat']
        
        print(f"Processed settings data: {update_data}")  # Debug log
        
        serializer = UserSettingsSerializer(request.user, data=update_data, partial=True)
        if serializer.is_valid():
            user = serializer.save()
            
            # Return updated settings in frontend format
            response_data = serializer.data
            response_data['settings'] = {
                'emailNotifications': getattr(user, 'email_notifications', True),
                'budgetAlerts': getattr(user, 'budget_alerts', True),
                'monthlyReports': getattr(user, 'monthly_reports', True),
                'reportFormat': getattr(user, 'preferred_currency', 'excel'),
                'twoFactorEnabled': False
            }
            
            return Response({
                'message': 'Settings updated successfully!',
                'settings': response_data,
                'user': response_data
            })
        else:
            print(f"Settings serializer errors: {serializer.errors}")  # Debug log
            return Response({
                'error': 'Validation failed',
                'details': serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
    
    return Response({'error': 'Method not allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

# =============================================================================
# TRANSACTION ENDPOINTS
# =============================================================================

class TransactionPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = 'page_size'
    max_page_size = 100

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def transactions_list(request: Request) -> Response:
    """List and create transactions"""
    if request.method == 'GET':
        queryset = Transaction.objects.filter(user=request.user)
        
        # Filtering
        transaction_type = request.query_params.get('type')
        if transaction_type:
            queryset = queryset.filter(type=transaction_type)
        
        category_id = request.query_params.get('category')
        if category_id:
            queryset = queryset.filter(category_id=category_id)
        
        start_date = request.query_params.get('start_date')
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        
        end_date = request.query_params.get('end_date')
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        
        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) | Q(description__icontains=search)
            )
        
        # Pagination
        paginator = TransactionPagination()
        page = paginator.paginate_queryset(queryset, request)
        if page is not None:
            serializer = TransactionSerializer(page, many=True)
            return paginator.get_paginated_response(serializer.data)
        
        serializer = TransactionSerializer(queryset, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        serializer = TransactionSerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            transaction = serializer.save(user=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    return Response({'error': 'Method not allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def transaction_detail(request: Request, transaction_id: int) -> Response:
    """Get, update, or delete a specific transaction"""
    try:
        transaction = Transaction.objects.get(id=transaction_id, user=request.user)
    except Transaction.DoesNotExist:
        return Response({'error': 'Transaction not found'}, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = TransactionSerializer(transaction)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        serializer = TransactionSerializer(transaction, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        transaction.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    return Response({'error': 'Method not allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

# =============================================================================
# CATEGORY ENDPOINTS
# =============================================================================

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def categories_list(request: Request) -> Response:
    """List and create categories"""
    if request.method == 'GET':
        queryset = Category.objects.filter(user=request.user)
        
        category_type = request.query_params.get('type')
        if category_type:
            queryset = queryset.filter(type=category_type)
        
        serializer = CategorySerializer(queryset, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        serializer = CategorySerializer(data=request.data, context={'request': request})
        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    return Response({'error': 'Method not allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

@api_view(['GET', 'PUT', 'DELETE'])
@permission_classes([IsAuthenticated])
def category_detail(request: Request, category_id: int) -> Response:
    """Get, update, or delete a specific category"""
    try:
        category = Category.objects.get(id=category_id, user=request.user)
    except Category.DoesNotExist:
        return Response({'error': 'Category not found'}, status=status.HTTP_404_NOT_FOUND)
    
    if request.method == 'GET':
        serializer = CategorySerializer(category)
        return Response(serializer.data)
    
    elif request.method == 'PUT':
        serializer = CategorySerializer(category, data=request.data, partial=True, context={'request': request})
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    elif request.method == 'DELETE':
        category.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    
    return Response({'error': 'Method not allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

# =============================================================================
# REPORTS ENDPOINTS
# =============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reports_monthly(request: Request) -> Response:
    """Get monthly financial reports"""
    try:
        # Get query parameters
        year = request.query_params.get('year', timezone.now().year)
        month = request.query_params.get('month', timezone.now().month)
        
        year = int(year)
        month = int(month)
        
        # Get or create monthly report
        report, created = MonthlyReport.objects.get_or_create(
            user=request.user,
            year=year,
            month=month,
            defaults={
                'total_income': 0,
                'total_expenses': 0,
                'net_income': 0,
                'income_by_category': {},
                'expenses_by_category': {},
                'compared_to_previous_month': {},
                'budget_analysis': {}
            }
        )
        
        if created or not report.is_final:
            # Calculate report data
            transactions = Transaction.objects.filter(
                user=request.user,
                date__year=year,
                date__month=month
            )
            
            income_transactions = transactions.filter(type='income')
            expense_transactions = transactions.filter(type='expense')
            
            report.total_income = income_transactions.aggregate(Sum('amount'))['amount__sum'] or 0
            report.total_expenses = expense_transactions.aggregate(Sum('amount'))['amount__sum'] or 0
            report.net_income = report.total_income - report.total_expenses
            
            # Category breakdown
            income_by_category = {}
            for transaction in income_transactions:
                category_name = transaction.category.name if transaction.category else 'Uncategorized'
                income_by_category[category_name] = income_by_category.get(category_name, 0) + float(transaction.amount)
            
            expenses_by_category = {}
            for transaction in expense_transactions:
                category_name = transaction.category.name if transaction.category else 'Uncategorized'
                expenses_by_category[category_name] = expenses_by_category.get(category_name, 0) + float(transaction.amount)
            
            report.income_by_category = income_by_category
            report.expenses_by_category = expenses_by_category
            report.save()
        
        serializer = MonthlyReportSerializer(report)
        return Response(serializer.data)
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def reports_download(request: Request, year: int, month: int) -> HttpResponse:
    """Download monthly report as CSV"""
    try:
        # Get report data
        report = MonthlyReport.objects.get(user=request.user, year=year, month=month)
        
        # Create CSV response
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="financial_report_{year}_{month}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Financial Report', f'{month}/{year}'])
        writer.writerow([])
        writer.writerow(['Summary'])
        writer.writerow(['Total Income', f'${report.total_income}'])
        writer.writerow(['Total Expenses', f'${report.total_expenses}'])
        writer.writerow(['Net Income', f'${report.net_income}'])
        writer.writerow([])
        
        writer.writerow(['Income by Category'])
        for category, amount in report.income_by_category.items():
            writer.writerow([category, f'${amount}'])
        
        writer.writerow([])
        writer.writerow(['Expenses by Category'])
        for category, amount in report.expenses_by_category.items():
            writer.writerow([category, f'${amount}'])
        
        return response
        
    except MonthlyReport.DoesNotExist:
        return Response({'error': 'Report not found'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

# =============================================================================
# BUDGET ENDPOINTS
# =============================================================================

@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
def budget_list(request: Request) -> Response:
    """List and create budgets"""
    if request.method == 'GET':
        budgets = Budget.objects.filter(user=request.user)
        serializer = BudgetSerializer(budgets, many=True)
        return Response(serializer.data)
    
    elif request.method == 'POST':
        serializer = BudgetSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(user=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    return Response({'error': 'Method not allowed'}, status=status.HTTP_405_METHOD_NOT_ALLOWED)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def budget_overview(request: Request) -> Response:
    """Get budget overview for the user"""
    try:
        current_month = timezone.now().month
        current_year = timezone.now().year
        
        # Get current month's budget
        try:
            current_budget = Budget.objects.get(
                user=request.user,
                year=current_year,
                month=current_month
            )
            serializer = BudgetSerializer(current_budget)
            budget_data = serializer.data
        except Budget.DoesNotExist:
            budget_data = None
        
        # Get recent alerts
        recent_alerts = BudgetAlert.objects.filter(
            user=request.user,
            is_read=False
        )[:3]
        
        alerts_data = BudgetAlertSerializer(recent_alerts, many=True).data
        
        return Response({
            'current_budget': budget_data,
            'recent_alerts': alerts_data,
            'month': current_month,
            'year': current_year
        })
        
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def budget_alerts(request: Request) -> Response:
    """Get budget alerts for the user"""
    alerts = BudgetAlert.objects.filter(user=request.user, is_read=False)
    serializer = BudgetAlertSerializer(alerts, many=True)
    return Response(serializer.data)

# =============================================================================
# USER ACTIVITY AND SUGGESTIONS ENDPOINTS
# =============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def users_activity(request: Request) -> Response:
    """Get user activity log"""
    limit = int(request.query_params.get('limit', 10))
    
    # Get recent transactions as activity items
    recent_transactions = Transaction.objects.filter(
        user=request.user
    ).order_by('-created_at')[:limit]
    
    activity_items = []
    for transaction in recent_transactions:
        activity_items.append({
            'id': f"transaction_{transaction.pk}",
            'type': 'transaction',
            'description': f"Added {transaction.type} transaction: {transaction.title}",
            'timestamp': transaction.created_at.isoformat(),
            'metadata': {
                'transaction_id': transaction.pk,
                'amount': float(transaction.amount),
                'type': transaction.type
            }
        })
    
    return Response(activity_items)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def transaction_suggestions(request: Request) -> Response:
    """Get transaction title suggestions based on user's history"""
    transaction_type = request.query_params.get('type', 'expense')
    query = request.query_params.get('q', '')
    
    # Get unique transaction titles from user's history
    transactions = Transaction.objects.filter(
        user=request.user,
        type=transaction_type
    ).exclude(title__isnull=True).exclude(title__exact='')
    
    if query:
        transactions = transactions.filter(
            Q(title__icontains=query) | Q(description__icontains=query)
        )
    
    # Get unique titles
    titles = list(set([
        t.title for t in transactions if t.title
    ] + [
        t.description for t in transactions if t.description and not t.title
    ]))
    
    # Limit results
    suggestions = sorted(titles)[:10]
    
    return Response({
        'suggestions': suggestions,
        'type': transaction_type,
        'query': query
    })

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def users_avatar(request: Request) -> Response:
    """Upload user avatar"""
    if 'avatar' not in request.FILES:
        return Response({'error': 'No avatar file provided'}, status=status.HTTP_400_BAD_REQUEST)
    
    avatar_file = request.FILES['avatar']
    
    # Simple file validation
    if avatar_file.size > 5 * 1024 * 1024:  # 5MB limit
        return Response({'error': 'File size too large'}, status=status.HTTP_400_BAD_REQUEST)
    
    if not avatar_file.content_type.startswith('image/'):
        return Response({'error': 'Invalid file type'}, status=status.HTTP_400_BAD_REQUEST)
    
    # Save avatar using the correct field name
    request.user.profile_picture = avatar_file
    request.user.save()
    
    # Return updated user data
    serializer = UserProfileSerializer(request.user)
    response_data = serializer.data
    response_data.update({
        'firstName': getattr(request.user, 'first_name', ''),
        'lastName': getattr(request.user, 'last_name', ''),
        'name': f"{getattr(request.user, 'first_name', '')} {getattr(request.user, 'last_name', '')}".strip(),
        'avatar': getattr(request.user, 'profile_picture', None),
        'settings': {
            'emailNotifications': getattr(request.user, 'email_notifications', True),
            'budgetAlerts': getattr(request.user, 'budget_alerts', True),
            'monthlyReports': getattr(request.user, 'monthly_reports', True),
            'reportFormat': getattr(request.user, 'preferred_currency', 'excel'),
            'twoFactorEnabled': False
        }
    })
    
    return Response({
        'message': 'Avatar uploaded successfully',
        'user': response_data
    })

@api_view(['DELETE'])
@permission_classes([IsAuthenticated])
def users_avatar_delete(request: Request) -> Response:
    """Delete user avatar"""
    # Clear the profile picture field
    request.user.profile_picture = None
    request.user.save()
    
    # Return updated user data
    serializer = UserProfileSerializer(request.user)
    response_data = serializer.data
    response_data.update({
        'firstName': getattr(request.user, 'first_name', ''),
        'lastName': getattr(request.user, 'last_name', ''),
        'name': f"{getattr(request.user, 'first_name', '')} {getattr(request.user, 'last_name', '')}".strip(),
        'avatar': None,
        'settings': {
            'emailNotifications': getattr(request.user, 'email_notifications', True),
            'budgetAlerts': getattr(request.user, 'budget_alerts', True),
            'monthlyReports': getattr(request.user, 'monthly_reports', True),
            'reportFormat': getattr(request.user, 'preferred_currency', 'excel'),
            'twoFactorEnabled': False
        }
    })
    
    return Response({
        'message': 'Avatar deleted successfully',
        'user': response_data
    })

# =============================================================================
# DASHBOARD STATISTICS ENDPOINTS
# =============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def dashboard_statistics(request: Request) -> Response:
    """Get dashboard statistics including month-over-month comparisons"""
    try:
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        current_date = timezone.now()
        current_month_start = current_date.replace(day=1)
        last_month_start = (current_month_start - timedelta(days=1)).replace(day=1)
        last_month_end = current_month_start - timedelta(days=1)
        
        # Current month transactions
        current_month_transactions = Transaction.objects.filter(
            user=request.user,
            date__gte=current_month_start
        )
        
        # Last month transactions
        last_month_transactions = Transaction.objects.filter(
            user=request.user,
            date__gte=last_month_start,
            date__lte=last_month_end
        )
        
        # Calculate current month totals
        current_income = current_month_transactions.filter(type='income').aggregate(
            total=Sum('amount')
        )['total'] or 0
        
        current_expenses = current_month_transactions.filter(type='expense').aggregate(
            total=Sum('amount')
        )['total'] or 0
        
        # Calculate last month totals
        last_income = last_month_transactions.filter(type='income').aggregate(
            total=Sum('amount')
        )['total'] or 0
        
        last_expenses = last_month_transactions.filter(type='expense').aggregate(
            total=Sum('amount')
        )['total'] or 0
        
        # Calculate percentage changes
        def calculate_percentage_change(current, previous):
            if previous == 0:
                return 100 if current > 0 else 0
            return ((current - previous) / previous) * 100
        
        income_change = calculate_percentage_change(float(current_income), float(last_income))
        expense_change = calculate_percentage_change(float(current_expenses), float(last_expenses))
        
        return Response({
            'current_month': {
                'income': float(current_income),
                'expenses': float(current_expenses),
                'balance': float(current_income) - float(current_expenses)
            },
            'last_month': {
                'income': float(last_income),
                'expenses': float(last_expenses),
                'balance': float(last_income) - float(last_expenses)
            },
            'changes': {
                'income_percentage': round(income_change, 1),
                'expense_percentage': round(expense_change, 1),
                'income_direction': 'up' if income_change >= 0 else 'down',
                'expense_direction': 'up' if expense_change >= 0 else 'down'
            },
            'transaction_counts': {
                'current_month': current_month_transactions.count(),
                'last_month': last_month_transactions.count(),
                'current_income_count': current_month_transactions.filter(type='income').count(),
                'current_expense_count': current_month_transactions.filter(type='expense').count()
            }
        })
        
    except Exception as e:
        return Response({
            'error': f'Statistics calculation failed: {str(e)}',
            'message': 'Failed to calculate dashboard statistics'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

# =============================================================================
# EXPORT ENDPOINTS
# =============================================================================

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def transactions_export(request: Request) -> HttpResponse:
    """Export transactions as Excel or PDF"""
    try:
        file_type = request.query_params.get('fileType', 'excel')
        
        # Get user's transactions
        transactions = Transaction.objects.filter(user=request.user).order_by('-date')
        
        if file_type.lower() == 'excel':
            # Create Excel file
            import openpyxl
            from django.http import HttpResponse
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Transactions"
            
            # Headers
            headers = ['Date', 'Title', 'Description', 'Type', 'Category', 'Amount']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Data
            for row, transaction in enumerate(transactions, 2):
                ws.cell(row=row, column=1, value=transaction.date.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=2, value=transaction.title or '')
                ws.cell(row=row, column=3, value=transaction.description or '')
                ws.cell(row=row, column=4, value=transaction.type.capitalize())
                ws.cell(row=row, column=5, value=transaction.category.name if transaction.category else 'Uncategorized')
                ws.cell(row=row, column=6, value=float(transaction.amount))
            
            # Save to BytesIO
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="transactions_export.xlsx"'
            return response
            
        elif file_type.lower() == 'pdf':
            # Create PDF file (using reportlab)
            try:
                from reportlab.lib.pagesizes import letter
                from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                from reportlab.lib.styles import getSampleStyleSheet
                from reportlab.lib import colors
                from io import BytesIO
                
                buffer = BytesIO()
                doc = SimpleDocTemplate(buffer, pagesize=letter)
                elements = []
                
                # Title
                styles = getSampleStyleSheet()
                title = Paragraph("Transaction Export", styles['Title'])
                elements.append(title)
                elements.append(Spacer(1, 12))
                
                # Table data
                data = [['Date', 'Title', 'Type', 'Category', 'Amount']]
                for transaction in transactions:
                    data.append([
                        transaction.date.strftime('%Y-%m-%d'),
                        transaction.title or '',
                        transaction.type.capitalize(),
                        transaction.category.name if transaction.category else 'Uncategorized',
                        f"${float(transaction.amount):.2f}"
                    ])
                
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                elements.append(table)
                doc.build(elements)
                
                buffer.seek(0)
                response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
                response['Content-Disposition'] = 'attachment; filename="transactions_export.pdf"'
                return response
                
            except ImportError:
                # Fallback to CSV if reportlab is not installed
                return Response({
                    'error': 'PDF generation not available. Please install reportlab package.',
                    'suggestion': 'Use Excel export instead.'
                }, status=status.HTTP_400_BAD_REQUEST)
        
        else:
            # Default to CSV
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="transactions_export.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Date', 'Title', 'Description', 'Type', 'Category', 'Amount'])
            
            for transaction in transactions:
                writer.writerow([
                    transaction.date.strftime('%Y-%m-%d'),
                    transaction.title or '',
                    transaction.description or '',
                    transaction.type.capitalize(),
                    transaction.category.name if transaction.category else 'Uncategorized',
                    f"{float(transaction.amount):.2f}"
                ])
            
            return response
            
    except Exception as e:
        return Response({
            'error': f'Export failed: {str(e)}',
            'message': 'Failed to export transactions'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)